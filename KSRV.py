import os
import subprocess
import re
import json
import base64
import uuid
import tkinter as tk
from tkinter import filedialog, messagebox
import requests
import time
from tenacity import retry, wait_fixed, stop_after_attempt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import sqlite3



# Path to SQLite database
DB_PATH = 'credentials.db'

# Create and initialize the SQLite database
def initialize_database():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Create table if not exists
    cursor.execute('''CREATE TABLE IF NOT EXISTS users (username TEXT PRIMARY KEY, password TEXT)''')

    # Default credentials
    default_users = {
        "huan.ntn": "ZYHTREWQAS",
        "tinh.dt": "ASDFGHJKLQW",
        "hiep.nhh": "BVCXZLKJMN",
        "loan.htt": "QWERTYUIOPM",
        "loan.ttt": "ZXCVBNMASD",
        "quan.hh": "JHGFDERTYU",
        "phat.nt": "RFGHYUJMNP",
        "phat.ht": "TGBNHYUJVC",
        "trung.nhn": "YHGTREWQAS",
        "long.nt": "OPASDFGHJK",
        "trung.ns": "QAZWSXCVBN",
        "phuc.nm": "QWERTYZXCV",
        "hai.tt": "DFGTHYJUIK",
        "huyen.dt": "ZXCVBNMASD",
        "thuy.ntp": "POIUYTREWM",
        "admin": "Admin123456!"
    }

    # Insert default users
    for username, password in default_users.items():
        cursor.execute("INSERT OR IGNORE INTO users (username, password) VALUES (?, ?)", (username, password))

    conn.commit()
    conn.close()

# Đường dẫn lưu ảnh và file thông tin
IMAGE_SAVE_PATH = 'cccd_images'
INFO_FILE_PATH = 'info.txt'
COOKIE_FILE_PATH = "cookie.txt"
TEMP_FILE_PATH = "card_no_temp.txt"
SERVER_URL = "http://localhost:81/enroll/updateFaceData.php"

# Tạo thư mục để lưu trữ hình ảnh nếu chưa tồn tại
if not os.path.exists(IMAGE_SAVE_PATH):
    os.makedirs(IMAGE_SAVE_PATH)

# Đọc cookie từ file
def read_cookie():
    with open(COOKIE_FILE_PATH, "r") as f:
        return f.read().strip()

# Đọc thông tin từ dòng JSON duy nhất trong file info.txt
def read_info():
    try:
        with open(INFO_FILE_PATH, "r", encoding="utf-8") as f:
            line = f.readline().strip()
            return json.loads(line) if line else {}
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"Error reading info.txt: {e}")
        return {}

# Hàm gửi yêu cầu POST và kiểm tra response
@retry(stop=stop_after_attempt(3), wait=wait_fixed(2))
def send_post_request(burp0_url, burp0_headers, burp0_cookies, burp0_data, file_path):
    response = requests.post(burp0_url, headers=burp0_headers, cookies=burp0_cookies, data=burp0_data)
    print(response.text)

    if not response.ok or response.json().get("success", True) is False:
        print(f"Request failed with status code {response.status_code}. Attempting to update cookies and retry...")
        # Chạy script auto.py để cập nhật lại thông tin cookie
        update_cookie()
        # Đọc lại cookie mới từ file sau khi auto.py đã cập nhật
        burp0_cookies["PHPSESSID"] = read_cookie()
        # Gửi lại yêu cầu POST
        response = requests.post(burp0_url, headers=burp0_headers, cookies=burp0_cookies, data=burp0_data)
        print(response.text)
        if response.ok and response.json().get("success", True) is True:
            # Xóa hình ảnh khỏi thư mục sau khi gửi thành công
            os.remove(file_path)
            print(f"Deleted file: {file_path}")
            # Xóa nội dung trong file info.txt
            with open(INFO_FILE_PATH, "w") as f:
                f.truncate()
            print("Cleared info.txt")
    else:
        print("Request successful!")
        # Xóa hình ảnh khỏi thư mục sau khi gửi thành công
        os.remove(file_path)
        print(f"Deleted file: {file_path}")
        # Xóa nội dung trong file info.txt
        with open(INFO_FILE_PATH, "w") as f:
            f.truncate()
        print("Cleared info.txt")
    return response.text

# Hàm để kiểm tra và lưu thông tin CCCD
def save_info(text):
    # Loại bỏ các ký tự "||||" dư thừa ở cuối chuỗi
    text = text.rstrip('|')
    
    new_format_match = re.match(r'(\d*)\|(\d*)\|([^|]*)\|(\d{8})?\|(Nam|Nữ)?\|([^|]*)\|(\d{8})?', text)
    if new_format_match:
        cccd_number = new_format_match.group(1)
        cmnd_number = new_format_match.group(2)
        name = new_format_match.group(3)
        dob = new_format_match.group(4)
        gender = new_format_match.group(5)
        address = new_format_match.group(6)
        issue_date = new_format_match.group(7)

        # Kiểm tra và chuyển đổi định dạng ngày sinh và ngày cấp CCCD nếu có
        dob = f"{dob[:2]}/{dob[2:4]}/{dob[4:]}" if dob else ''
        issue_date = f"{issue_date[:2]}/{issue_date[2:4]}/{issue_date[4:]}" if issue_date else ''

        info = {
            'Số CCCD': cccd_number if cccd_number else '',
            'Số CMND': cmnd_number if cmnd_number else '',
            'Họ và tên': name if name else '',
            'Giới tính': gender if gender else '',
            'Ngày sinh': dob if dob else '',
            'Nơi thường trú': address if address else '',
            'Ngày cấp CCCD': issue_date if issue_date else ''
        }

    else:
        cccd_number_match = re.search(r'Số CCCD:\s*(\d+)', text)
        cmnd_number_match = re.search(r'Số CMND:\s*(\d+)', text)
        name_match = re.search(r'Họ và tên:\s*(.+)', text)
        gender_match = re.search(r'Giới tính:\s*(Nam|Nữ)', text)
        dob_match = re.search(r'Ngày sinh:\s*(\d{2}/\d{2}/\d{4})', text)
        address_match = re.search(r'Nơi thường trú:\s*(.+)', text)
        issue_date_match = re.search(r'Ngày cấp CCCD:\s*(\d{2}/\d{2}/\d{4})', text)

        if cccd_number_match:
            cccd_number = cccd_number_match.group(1)
            info = {
                'Số CCCD': cccd_number,
                'Số CMND': cmnd_number_match.group(1) if cmnd_number_match else '',
                'Họ và tên': name_match.group(1) if name_match else '',
                'Giới tính': gender_match.group(1) if gender_match else '',
                'Ngày sinh': dob_match.group(1) if dob_match else '',
                'Nơi thường trú': address_match.group(1) if address_match else '',
                'Ngày cấp CCCD': issue_date_match.group(1) if issue_date_match else ''
            }
        else:
            return None

    with open(INFO_FILE_PATH, 'w', encoding='utf-8') as f:
        json.dump(info, f, ensure_ascii=False)
        f.write('\n')
    return info


# Hàm để chọn ảnh khuôn mặt
def browse_image():
    file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
    if file_path:
        image_path_entry.delete(0, tk.END)
        image_path_entry.insert(0, file_path)
# Dictionary để ánh xạ lựa chọn giao diện và giá trị group_id
group_id_mapping = {
    "0. Cán bộ C06": "0",
    "1. Unauthorized": "1",
    "2. Khách": "2"
}
# Function to save registered CCCD information with timestamp and username to a log file
def log_cccd_info(info, username):
    file_name = "Log_CCCD.xlsx"

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "CCCD Logs"

        # Setting headers
        headers = ["Tài khoản trực ban", "Thời gian tải lên hồ sơ", "Số CCCD", "Số CMND", "Họ và tên", "Giới tính", "Ngày sinh", "Nơi thường trú", "Ngày cấp CCCD"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header

    # Append the new log entry
    row_data = [
        username,
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        info.get('Số CCCD', 'N/A'),
        info.get('Số CMND', 'N/A'),
        info.get('Họ và tên', 'N/A'),
        info.get('Giới tính', 'N/A'),
        info.get('Ngày sinh', 'N/A'),
        info.get('Nơi thường trú', 'N/A'),
        info.get('Ngày cấp CCCD', 'N/A')
    ]
    ws.append(row_data)

    # Save workbook
    try:
        wb.save(file_name)
        # messagebox.showinfo("Saved", f"Information logged and saved to {file_name}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save log to Excel file: {e}")

# Update the upload_info function to log successful registrations
def upload_info():
    cccd_info = cccd_info_entry.get()
    image_path = image_path_entry.get()
    group_id_display = group_id_var.get()
    group_id = group_id_mapping.get(group_id_display)
    
    if cccd_info and image_path:
        info = save_info(cccd_info)
        if info:
            face_uuid = str(uuid.uuid4())
            burp0_headers = {
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "X-Requested-With": "XMLHttpRequest",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36",
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "Origin": "http://localhost:81:81",
                "Sec-Fetch-Site": "same-origin",
                "Sec-Fetch-Mode": "cors",
                "Sec-Fetch-Dest": "empty",
                "Referer": "http://localhost:81:81/enroll/Enroll.php",
                "Accept-Encoding": "gzip, deflate",
                "Accept-Language": "en-US,en;q=0.9",
                "Connection": "close"
            }
            burp0_cookies = {"PHPSESSID": read_cookie()}

            with open(image_path, "rb") as image_file:
                encoded_image_data = base64.b64encode(image_file.read()).decode('utf-8')

            if not os.path.exists(TEMP_FILE_PATH):
                with open(TEMP_FILE_PATH, "w") as temp_file:
                    temp_file.write("1")

            with open(TEMP_FILE_PATH, "r") as temp_file:
                current_card_no = int(temp_file.read())

            next_card_no = current_card_no + 1

            with open(TEMP_FILE_PATH, "w") as temp_file:
                temp_file.write(str(next_card_no).zfill(6))

            burp0_data = {
                "face_uuid": face_uuid,
                "id": info.get("Số CCCD", ''),
                "id2": info.get("Giới tính", ''),
                "name": info.get("Họ và tên", ''),
                "email": '',
                "wiegand_b64": '',
                "card_no": f"001-{str(next_card_no).zfill(5)}",
                "card_bits": "26",
                "group_id": group_id,
                "new_face": "1",
                "person_img": '',
                "img": f'["data:image/jpeg;base64,{encoded_image_data}"]'
            }

            response = send_post_request(SERVER_URL, burp0_headers, burp0_cookies, burp0_data, image_path)

            pattern = r'"status":\s*\[(\d+)\]'
            match = re.search(pattern, response)
            if match:
                status_value = int(match.group(1))
                if status_value == 0:
                    # Log the information on successful upload including username and timestamp
                    log_cccd_info(info, logged_in_user)  # Assuming logged_in_user stores the username of the logged account
                    messagebox.showinfo("Success", "Thông tin đã được tải lên thành công!")                    
                    # messagebox.showinfo("Success", "Upload thành công và thông tin đã được lưu vào log_access.txt!")
                else:
                    messagebox.showerror("Failed", "Upload thất bại! Vui lòng thử lại.")
                    with open(TEMP_FILE_PATH, "w") as temp_file:
                        temp_file.write(str(current_card_no).zfill(6))
            else:
                messagebox.showerror("Failed", "Không tìm thấy pattern 'status' trong response")
                with open(TEMP_FILE_PATH, "w") as temp_file:
                    temp_file.write(str(current_card_no).zfill(6))
        else:
            messagebox.showerror("Error", "Thông tin CCCD không hợp lệ.")
    else:
        messagebox.showerror("Error", "Vui lòng nhập đầy đủ thông tin và chọn ảnh.")


# Hàm để cập nhật cookie
def update_cookie():
    driver = webdriver.Chrome()  # or webdriver.Firefox()
    try:
        driver.get("http://localhost:81/login.php")
        username = "admin"
        password = "admin"
        input_username = driver.find_element(By.ID, "uname")
        input_password = driver.find_element(By.ID, "pwd")
        input_username.send_keys(username)
        input_password.send_keys(password)
        input_password.send_keys(Keys.ENTER)
        # time.sleep(5)
        cookies = driver.get_cookies()
        with open(COOKIE_FILE_PATH, "w") as cookie_file:
            for cookie in cookies:
                if cookie['name'] == 'PHPSESSID':
                    cookie_file.write(cookie['value'] + "\n")
                    break
    finally:
        driver.quit()

# Hàm để thực hiện đăng nhập và cập nhật trạng thái
# def login():
    # update_cookie()
    # status = "Đăng nhập thành công!" if os.path.exists(COOKIE_FILE_PATH) else "Lỗi đăng nhập!"
    # login_status_var.set(status)
    
def login():
    update_cookie()
    status = "Đăng nhập thành công!" if os.path.exists(COOKIE_FILE_PATH) else "Lỗi đăng nhập!"
    login_status_var.set(status)
    # Enable buttons if login is successful
    if os.path.exists(COOKIE_FILE_PATH):
        upload_button.config(state="normal")
    else:
        upload_button.config(state="disabled")
            


# Hàm để cập nhật màu sắc của dropdown dựa trên lựa chọn
def update_dropdown_color(*args):
    selected_option = group_id_var.get()
    if selected_option == "0. Khách":
        group_id_dropdown.config(bg="#00BFFF", fg="white")  # Màu xanh sáng
        group_id_dropdown.config(font=("Helvetica", 12, "bold"))
    elif selected_option == "1. Unauthorized":
        group_id_dropdown.config(bg="#333333", fg="white")  # Màu đen xám
        group_id_dropdown.config(font=("Helvetica", 12, "bold"))
    elif selected_option == "2. Cán bộ C06":
        group_id_dropdown.config(bg="#006400", fg="white")  # Màu xanh lá đậm
        group_id_dropdown.config(font=("Helvetica", 12, "bold"))
    else:
        group_id_dropdown.config(bg="SystemButtonFace", fg="black")
        group_id_dropdown.config(font=("Helvetica", 12))
def open_camera_app():
    try:
        # Gọi đến ứng dụng camera mặc định của Windows
        subprocess.run("start microsoft.windows.camera:", shell=True, check=True)
    except subprocess.CalledProcessError as e:
        print(f"Failed to open camera app: {e}")
        
# Hàm để gọi phần mềm "KSRV_v2_Batch"
def open_ksrv_v2_batch():
    try:
        subprocess.Popen(["KSRV_v2_Batch"])  # Gọi shortcut của phần mềm
    except Exception as e:
        messagebox.showerror("Error", f"Không thể mở KSRV_v2_Batch: {e}")

# Hàm để gọi phần mềm "Delete_Batch"
def open_delete_batch():
    try:
        subprocess.Popen(["Delete_Batch"])  # Gọi shortcut của phần mềm
    except Exception as e:
        messagebox.showerror("Error", f"Không thể mở Delete_Batch: {e}")    

        
# Khởi tạo dữ liệu
# user_data = {
    # "admin": "admin",
    # "trucban": "trucban"
# }
# user_data = {
        # "huan.ntn": "ZYHTREWQAS",    #Nguyễn Thị Ngọc Huấn
        # "tinh.dt": "ASDFGHJKLQW",    #Đặng Thị Tình
        # "hiep.nhh": "BVCXZLKJMN",    #Nguyễn Hồ Hoàng Hiệp
        # "loan.htt": "QWERTYUIOPM",   #Hồ Thị Thanh Loan
        # "loan.ttt": "ZXCVBNMASD",    #Trần Thị Thanh Loan 
        # "quan.hh": "JHGFDERTYU",     #Hoàng Hồng Quân
        # "phat.nt": "RFGHYUJMNP",     #Nguyễn Tiến Phát
        # "phat.ht": "TGBNHYUJVC",     #Hoàng Tấn Phát
        # "trung.nhn": "YHGTREWQAS",   #Nguyễn Hữu Ngọc Trung
        # "long.nt": "OPASDFGHJK",     #Nguyễn Trọng Long
        # "trung.ns": "QAZWSXCVBN",    #Nguyễn Sinh Trung
        # "phuc.nm": "QWERTYZXCV",     #Nguyễn Minh Phúc
        # "hai.tt": "DFGTHYJUIK",      #Trần Thanh Hải
        # "huyen.dt": "ZXCVBNMASD",    #Đỗ Thị Huyền
        # "thuy.ntp": "POIUYTREWM",    #Nguyễn Thị Phương Thuý
        # "admin": "Admin123456!"      
# }

logged_in_user = None
shift_data = {}
       
# def authenticate_user():
    # global logged_in_user
    # username = username_entry.get()
    # password = password_entry.get()
    
    # Verify the username and password against user_data
    # if username in user_data and user_data[username] == password:
        # logged_in_user = username
        # check_in()
        # login_window.destroy()
        # create_gui()  # Open the main GUI interface
    # else:
        # messagebox.showerror("Error", "Tên tài khoản hoặc mật khẩu không đúng.")
        
# Authenticate user
def authenticate_user():
    global logged_in_user
    username = username_entry.get()
    password = password_entry.get()

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM users WHERE username=? AND password=?", (username, password))
    result = cursor.fetchone()
    
    conn.close()

    if result:
        logged_in_user = username
        check_in()  # Implement your check_in logic
        login_window.destroy()
        create_gui()  # Open the main GUI
    else:
        messagebox.showerror("Error", "Tên tài khoản hoặc mật khẩu không đúng.")        

# Hàm check-in
def check_in():
    global logged_in_user
    if logged_in_user:
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        shift_data[logged_in_user] = [current_time]
    else:
        messagebox.showerror("Error", "Bạn phải đăng nhập trước")

# Hàm check-out
def check_out():
    global logged_in_user
    if logged_in_user:
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if logged_in_user in shift_data and len(shift_data[logged_in_user]) == 1:
            shift_data[logged_in_user].append(current_time)
            save_to_excel()
        else:
            messagebox.showerror("Error", "Bạn phải đăng nhập trước")
    else:
        messagebox.showerror("Error", "Bạn phải đăng nhập trước")

# Hàm lưu vào file Excel
def save_to_excel():
    current_date = datetime.now().strftime("%d-%m-%Y")
    file_name = f"{current_date}.xlsx"
    
    # Kiểm tra xem file có tồn tại không
    if os.path.exists(file_name):
        # Mở workbook đã tồn tại
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        # Tạo workbook mới nếu file không tồn tại
        wb = Workbook()
        ws = wb.active
        ws.title = "Thoi gian ca truc"
        
        # Tạo tiêu đề cột
        headers = ["Employee", "Check-In Time", "Check-Out Time"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header

    # Lưu dữ liệu check-in/check-out
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        if cell.value in shift_data:
            # Cập nhật dữ liệu nếu nhân viên đã có trong file
            row_num = cell.row
            ws[f"B{row_num}"] = shift_data[cell.value][0]
            ws[f"C{row_num}"] = shift_data[cell.value][1]
            del shift_data[cell.value]
    
    # Thêm các nhân viên chưa có trong file
    for employee, times in shift_data.items():
        row_num = ws.max_row + 1
        ws[f"A{row_num}"] = employee
        ws[f"B{row_num}"] = times[0]
        ws[f"C{row_num}"] = times[1]
    
    # Lưu workbook
    wb.save(file_name)
    messagebox.showinfo("Saved", f"Thông tin ca trực đã được lưu trong tệp {file_name}")


# Hàm để tạo cửa sổ đăng nhập
def create_login_window():
    global username_entry, password_entry, login_window, status_label
    login_window = tk.Tk()  # Khởi tạo biến toàn cục
    login_window.iconbitmap("favicon.ico")
    login_window.title("Đăng nhập")
    login_window.geometry("285x150")
    login_window.configure(bg="#F0F8FF")  # Thiết lập màu nền cho cửa sổ

    # Tạo frame cho giao diện đăng nhập
    login_frame = tk.Frame(login_window, bg="#F0F8FF")
    login_frame.pack(pady=10, padx=10, fill="x")
    
    button_frame = tk.Frame(login_frame)
    button_frame.grid(row=2, columnspan=2)
    # Tạo các thành phần giao diện đăng nhập
    tk.Label(login_frame, text="Tài khoản:", bg="#F0F8FF", width=10, font=("Helvetica", 11, "bold")).grid(row=0, column=0, padx=8, pady=8, sticky="e")
    username_entry = tk.Entry(login_frame)
    username_entry.grid(row=0, column=1, padx=10, pady=10)
    
    tk.Label(login_frame, text="Mật khẩu:", bg="#F0F8FF", width=10, font=("Helvetica", 11, "bold")).grid(row=1, column=0, padx=8, pady=8, sticky="e")
    password_entry = tk.Entry(login_frame, show="*")
    password_entry.grid(row=1, column=1, padx=10, pady=10)
    
    # tk.Button(login_frame, text="Đăng nhập", command=authenticate_user, bg="#32CD32", fg="white", font=("Helvetica", 12, "bold"), width=9).grid(row=3, columnspan=3, pady=6)
    # tk.Button(button_frame, text="Đăng nhập", command=authenticate_user, bg="#32CD32", fg="white", font=("Helvetica", 12, "bold"), width=11).pack(pady=3)
    # Nút đăng nhập
    login_button = tk.Button(button_frame, text="Đăng nhập", command=authenticate_user, bg="#32CD32", fg="white", font=("Helvetica", 12, "bold"), width=11)
    login_button.pack(pady=3)
    
    status_label = tk.Label(login_window, text="", bg="#F0F8FF")
    status_label.pack(pady=10)
    
    # Gán sự kiện nhấn phím Enter để gọi hàm authenticate_user
    login_window.bind('<Return>', lambda event: login_button.invoke())
    
    login_window.mainloop()
    
# Function to change the user's password
# def change_password():
    # def update_password():
        # current_password = current_password_entry.get()
        # new_password = new_password_entry.get()
        # confirm_password = confirm_password_entry.get()

        # if user_data.get(logged_in_user) != current_password:
            # messagebox.showerror("Error", "Mật khẩu hiện tại không đúng.")
        # elif new_password != confirm_password:
            # messagebox.showerror("Error", "Mật khẩu mới không khớp.")
        # else:
            # user_data[logged_in_user] = new_password
            # messagebox.showinfo("Success", "Mật khẩu đã được thay đổi thành công.")
            # change_password_window.destroy()

    # change_password_window = tk.Toplevel()
    # change_password_window.title("Đổi Mật Khẩu")
    # change_password_window.geometry("300x200")

    # tk.Label(change_password_window, text="Mật khẩu hiện tại:").pack(pady=5)
    # current_password_entry = tk.Entry(change_password_window, show="*")
    # current_password_entry.pack(pady=5)

    # tk.Label(change_password_window, text="Mật khẩu mới:").pack(pady=5)
    # new_password_entry = tk.Entry(change_password_window, show="*")
    # new_password_entry.pack(pady=5)

    # tk.Label(change_password_window, text="Xác nhận mật khẩu:").pack(pady=5)
    # confirm_password_entry = tk.Entry(change_password_window, show="*")
    # confirm_password_entry.pack(pady=5)

    # tk.Button(change_password_window, text="Cập nhật", command=update_password).pack(pady=10)
    
    
# Change password function
def change_password():
    def update_password():
        current_password = current_password_entry.get()
        new_password = new_password_entry.get()
        confirm_password = confirm_password_entry.get()

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute("SELECT password FROM users WHERE username=?", (logged_in_user,))
        result = cursor.fetchone()

        if result and result[0] == current_password:
            if new_password == confirm_password:
                cursor.execute("UPDATE users SET password=? WHERE username=?", (new_password, logged_in_user))
                conn.commit()
                messagebox.showinfo("Success", "Mật khẩu đã được thay đổi thành công.")
                change_password_window.destroy()
            else:
                messagebox.showerror("Error", "Mật khẩu mới không khớp.")
        else:
            messagebox.showerror("Error", "Mật khẩu hiện tại không đúng.")
        
        conn.close()

    change_password_window = tk.Toplevel()
    change_password_window.title("Đổi Mật Khẩu")
    change_password_window.geometry("300x255")
    change_password_window.iconbitmap("favicon.ico")
    change_password_window.configure(bg="#F0F8FF")

   # status_label = tk.Label(
        # frame4,
        # textvariable=login_status_var,
        # font=("Helvetica", 12, "italic"),  # Set font to italic
        # fg="#FF5733",  # Set text color to a custom color (e.g., orange)
        # bg="#F0F8FF"
    # )
    
    tk.Label(change_password_window, text="Mật khẩu hiện tại:", font=("Helvetica", 10, "bold"), bg="#F0F8FF").pack(pady=5)
    current_password_entry = tk.Entry(change_password_window, show="*")
    current_password_entry.pack(pady=5)

    tk.Label(change_password_window, text="Mật khẩu mới:", font=("Helvetica", 10, "bold"), bg="#F0F8FF").pack(pady=5)
    new_password_entry = tk.Entry(change_password_window, show="*")
    new_password_entry.pack(pady=5)

    tk.Label(change_password_window, text="Xác nhận mật khẩu:", font=("Helvetica", 10, "bold"), bg="#F0F8FF").pack(pady=5)
    confirm_password_entry = tk.Entry(change_password_window, show="*")
    confirm_password_entry.pack(pady=5)

    tk.Button(change_password_window, text="Cập nhật", command=update_password, bg="#4682B4", fg="white", font=("Helvetica", 12, "bold"), width=8).pack(pady=10)

# Hàm chính để tạo giao diện
def create_gui():
    global cccd_info_entry, image_path_entry, login_status_var, group_id_var, group_id_dropdown, upload_button

    root = tk.Tk()
    root.title("ĐĂNG KÝ RA VÀO")
    root.geometry("824x355")  # Thay đổi kích thước cửa sổ
    
    # Khởi tạo biến lưu trữ giá trị lựa chọn của dropdown
    group_id_var = tk.StringVar(value="0. Khách")
    
    # Thiết lập màu nền và font chữ
    root.configure(bg="#F0F8FF")
    font_bold = ("Helvetica", 14, "bold")
    
    # Đặt biểu tượng ứng dụng bằng hình ảnh .ico
    root.iconbitmap("favicon.ico")

    # Tạo các frame cho từng hàng
    frame1 = tk.Frame(root, bg="#F0F8FF")
    frame1.pack(pady=10, padx=10, fill="x")

    frame2 = tk.Frame(root, bg="#F0F8FF")
    frame2.pack(pady=10, padx=10, fill="x")

    frame7 = tk.Frame(root, bg="#F0F8FF")
    frame7.pack(pady=10, padx=10, fill="x")    
    
    frame6 = tk.Frame(root, bg="#F0F8FF")
    frame6.pack(pady=10, padx=10, fill="x")        

    frame5 = tk.Frame(root, bg="#F0F8FF")  # Di chuyển frame5 lên trước frame3
    frame5.pack(pady=10, padx=10, fill="x")
    
    frame3 = tk.Frame(root, bg="#F0F8FF")
    frame3.pack(pady=10, padx=10, fill="x")
    

    frame4 = tk.Frame(root, bg="#F0F8FF")
    frame4.pack(pady=10, padx=10, fill="x")
    
    # Thêm nút đổi mật khẩu
    frame8 = tk.Frame(root, bg="#F0F8FF")
    frame8.pack(pady=10, padx=10, fill="x")
    
    tk.Label(frame1, text="Thông tin CCCD:", font=font_bold, bg="#F0F8FF").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    cccd_info_entry = tk.Entry(frame1, width=50, font=("Helvetica", 12))
    cccd_info_entry.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(frame2, text="Đường dẫn ảnh:", font=font_bold, bg="#F0F8FF").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    image_path_entry = tk.Entry(frame2, width=50, font=("Helvetica", 12))
    image_path_entry.grid(row=0, column=1, padx=5, pady=5)
    tk.Button(frame2, text="Chọn ảnh", command=browse_image, bg="#4682B4", fg="white", font=("Helvetica", 12, "bold"), width=15).grid(row=0, column=2, padx=5, pady=5)
    # Nút mở phần mềm KSRV_v2_Batch
    tk.Button(frame6, text="Tạo Profiles theo lô", command=open_ksrv_v2_batch, bg="#DC1400", fg="white", font=("Helvetica", 12, "bold"), width=20).pack(side="left", padx=5)    
    # Nút xóa profile theo lô
    tk.Button(frame6, text="Xóa Profiles theo lô", command=open_delete_batch, bg="#0D0D0D", fg="white", font=("Helvetica", 12, "bold"), width=20).pack(side="left", padx=5) 
    tk.Button(frame6, text="Đổi Mật Khẩu", command=change_password, bg="#FF5E93", fg="white", font=("Helvetica", 12, "bold"), width=20).pack(side="left", padx=5)    
    tk.Button(frame6, text="Mở Camera", command=open_camera_app, bg="#FFA500",fg="white", font=("Helvetica", 12, "bold"), width=15).pack(side="right", padx=5)

    
    tk.Label(frame5, text="Chọn Group:", font=font_bold, bg="#F0F8FF").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    group_id_var = tk.StringVar(value="2. Khách")
    group_id_dropdown = tk.OptionMenu(frame5, group_id_var, "0. Cán bộ C06", "1. Unauthorized", "2. Khách")
    group_id_dropdown.config(font=("Helvetica", 12))
    group_id_dropdown.grid(row=0, column=1, padx=5, pady=5)
    # Cập nhật màu sắc ngay khi có sự thay đổi lựa chọn
    group_id_var.trace("w", update_dropdown_color)       

    # tk.Button(frame3, text="Tải lên thông tin", command=upload_info, bg="#32CD32", fg="white", font=("Helvetica", 12, "bold"), width=20).pack(side="left", padx=5)
    upload_button = tk.Button(frame3, text="Tải lên thông tin", command=upload_info, state="disabled", font=("Helvetica", 12, "bold"), bg="#32CD32", fg="white")
    upload_button.pack(side="left", padx=5)
    
    tk.Button(frame3, text="Đăng nhập", command=login, bg="#FFD700", fg="black", font=("Helvetica", 12, "bold"), width=15).pack(side="right", padx=5)
    # Button to change password
    # tk.Button(frame8, text="Đổi Mật Khẩu", command=change_password, bg="#FFA500", fg="white", font=("Helvetica", 12, "bold"), width=18).pack(side="right", padx=5)    
    login_status_var = tk.StringVar()
    # login_status_var.set("Hãy nhấn nút đăng nhập, để tiếp tục tải lên thông tin")
    # tk.Label(frame4, textvariable=login_status_var, font=font_bold, bg="#F0F8FF").pack()
    login_status_var.set("Hãy nhấn nút đăng nhập, để tiếp tục tải lên thông tin")

    # Label to display the login status with italicized text and custom color
    status_label = tk.Label(
        frame4,
        textvariable=login_status_var,
        font=("Helvetica", 12, "italic"),  # Set font to italic
        fg="#FF5733",  # Set text color to a custom color (e.g., orange)
        bg="#F0F8FF"
    )
    status_label.pack()
    
    # Cài đặt sự kiện để gọi hàm check-out khi đóng ứng dụng
    root.protocol("WM_DELETE_WINDOW", lambda: (check_out(), root.destroy()))   

    root.mainloop()

if __name__ == "__main__":
    initialize_database()
    create_login_window()

